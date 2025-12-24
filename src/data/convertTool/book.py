from openpyxl import load_workbook
import json
import os
from datetime import datetime

# ===== paths =====
EXCEL_PATH = "/mnt/data/TIPC選書.xlsx"
EXISTING_JSON_PATH = "/mnt/data/existing_books.json"   # ← change if needed
OUTPUT_JSON_PATH = "/mnt/data/TIPC_books_from_62.json"

# ===== load excel =====
wb = load_workbook(EXCEL_PATH)
ws = wb.active
headers = [cell.value for cell in ws[1]]

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    record = dict(zip(headers, row))
    if record.get("書本名"):
        rows.append(record)

# ===== start from specific book =====
START_TITLE = "鎏金舞台：你不可不知道的歌劇發展社會史"
start_index = next(i for i, r in enumerate(rows) if r.get("書本名") == START_TITLE)
rows = rows[start_index:]

# ===== helpers =====
def clean_people_field(value):
    if not value:
        return []
    return [
        p.strip()
        for p in str(value)
            .replace("\n", "")
            .replace("\r", "")
            .split("、")
        if p.strip()
    ]

today = datetime.now().strftime("%Y/%m/%d")

# ===== load existing json =====
existing_book_names = set()

if os.path.exists(EXISTING_JSON_PATH):
    with open(EXISTING_JSON_PATH, "r", encoding="utf-8") as f:
        existing_books = json.load(f)
        existing_book_names = {
            b.get("bookName") for b in existing_books if b.get("bookName")
        }

# ===== convert =====
result = []
duplicates = []

current_id = 62

for book in rows:
    book_name = book.get("書本名", "").strip()

    # skip duplicates
    if book_name in existing_book_names:
        duplicates.append(book_name)
        continue

    result.append({
        "id": str(current_id),
        "bookName": book_name,
        "author": clean_people_field(book.get("作者")),
        "image": f"/images/books/{book_name}.jpg",
        "uploadDate": today,
        "publisher": book.get("出版社", ""),
        "isbn": str(book.get("ISBN ")).replace(".0", "") if book.get("ISBN ") else ""
    })

    current_id += 1

# ===== output json =====
with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

# ===== terminal report =====
if duplicates:
    print("\n🚨 Duplicate books skipped:")
    for name in duplicates:
        print(f" - {name}")
else:
    print("\n✅ No duplicate books found.")

print(f"\n✅ JSON generated: {OUTPUT_JSON_PATH}")
print(f"📚 New books added: {len(result)}")
